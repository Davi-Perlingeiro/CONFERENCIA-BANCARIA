import streamlit as st
import pandas as pd
import glob
import os
import re
import unicodedata
import pdfplumber
from io import BytesIO
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title='Comparador Folha x Banco', layout='wide')


# ============================================================
# FUNCOES UTILITARIAS
# ============================================================

def normalizar_nome(nome):
    if pd.isna(nome):
        return ''
    nome = str(nome).strip().upper()
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(c for c in nome if unicodedata.category(c) != 'Mn')
    nome = re.sub(r'\s+', ' ', nome)
    nome = re.sub(r'^\d{2}\.\d{3}\.\d{3}\s+', '', nome)
    return nome


def extrair_digitos_visiveis_cpf(cpf_mascarado):
    if pd.isna(cpf_mascarado):
        return ''
    return ''.join(re.findall(r'\d+', str(cpf_mascarado)))


def cpf_limpo(cpf):
    if pd.isna(cpf):
        return ''
    return re.sub(r'[.\-\s]', '', str(cpf).strip())


def sugerir_origem(nome_arquivo):
    nome = nome_arquivo.upper()
    if 'RPA' in nome or 'PAGA' in nome:
        return 'RPA'
    if 'QUINZ' in nome:
        m = re.search(r'(\d{4})', nome_arquivo)
        cod = m.group(1) if m else ''
        local_m = re.search(r'PG_(\w+)', nome, re.IGNORECASE)
        local = local_m.group(1) if local_m else ''
        if local == 'QUINZ':
            after = re.search(r'QUINZ[_ -]+(?:OP\s+)?(\w+)', nome)
            local = after.group(1) if after else local
        return f'QUINZ {local} {cod}'.strip()
    if 'LIQUIDO' in nome or 'FOLHA' in nome or 'INTERM' in nome:
        m = re.search(r'(\d{4})', nome_arquivo)
        cod = m.group(1) if m else ''
        local_m = re.search(r'(?:PG_|INTERM[_ ]*)(\w+)', nome)
        local = local_m.group(1) if local_m else 'CLT'
        return f'{local} {cod}'.strip()
    return 'FOLHA'


# ============================================================
# LEITURA DE ARQUIVOS
# ============================================================

def ler_banco_bytes(file_bytes, nome_arquivo):
    df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
    header_row = None
    for i in range(min(30, len(df_raw))):
        row_str = ' '.join(str(v).lower() for v in df_raw.iloc[i].values if pd.notna(v))
        if 'favorecido' in row_str or 'benefici' in row_str:
            header_row = i
            break
    if header_row is None:
        header_row = 18
    df_banco = df_raw.iloc[header_row + 1:].copy()
    ncols = len(df_banco.columns)
    col_names = ['nome', 'cpf_cnpj', 'tipo_pag', 'ref_empresa', 'data_pag', 'valor', 'status']
    if ncols >= 7:
        df_banco.columns = col_names + [f'extra_{i}' for i in range(ncols - 7)]
    df_banco = df_banco[col_names[:min(ncols, 7)]]
    df_banco = df_banco.dropna(subset=['nome'])
    df_banco = df_banco[df_banco['nome'].astype(str).str.strip() != '']
    df_banco = df_banco[df_banco['nome'] != 'Total:']
    df_banco = df_banco[~df_banco['nome'].astype(str).str.lower().str.contains('favorecido')]
    df_banco['valor'] = pd.to_numeric(df_banco['valor'], errors='coerce')
    df_banco['nome_norm'] = df_banco['nome'].apply(normalizar_nome)
    df_banco['cpf_digitos'] = df_banco['cpf_cnpj'].apply(extrair_digitos_visiveis_cpf)
    return df_banco


def ler_folha_clt_bytes(file_bytes, origem):
    df = pd.read_excel(BytesIO(file_bytes), header=None)
    dados = df.iloc[8:].copy()
    dados = dados[dados[0].notna()]
    dados = dados[~dados[0].astype(str).str.contains('Total|Dpto|TOTAL|Resumo', na=False)]
    registros = []
    for _, row in dados.iterrows():
        nome = str(row[0]).strip()
        valor = row[15] if pd.notna(row[15]) else (row[16] if len(row) > 16 and pd.notna(row[16]) else None)
        if nome and valor is not None:
            try:
                registros.append({'nome': nome, 'nome_norm': normalizar_nome(nome),
                    'valor': float(valor), 'origem': origem, 'tipo': 'CLT', 'cpf': ''})
            except (ValueError, TypeError):
                pass
    return registros


def ler_folha_rpa_excel_bytes(file_bytes, origem='RPA'):
    df = pd.read_excel(BytesIO(file_bytes), header=None)
    dados = df.iloc[6:].copy()
    dados = dados[dados[2].notna()]
    dados = dados[~dados[2].astype(str).str.contains('Total|Nome|TOTAL', na=False)]
    registros = []
    for _, row in dados.iterrows():
        nome = str(row[2]).strip()
        valor = row[3]
        cpf = str(row[10]).strip() if len(row) > 10 and pd.notna(row[10]) else ''
        if nome and pd.notna(valor):
            try:
                registros.append({'nome': nome, 'nome_norm': normalizar_nome(nome),
                    'valor': float(valor), 'origem': origem, 'tipo': 'RPA', 'cpf': cpf_limpo(cpf)})
            except (ValueError, TypeError):
                pass
    return registros


def ler_folha_pdf_bytes(file_bytes, origem='PDF'):
    pdf = pdfplumber.open(BytesIO(file_bytes))
    registros = []
    current_nome = None
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue
        for line in text.split('\n'):
            m = re.match(r'(\d{6})\s+(.+?)\s+\d+[,.]', line)
            if m:
                current_nome = re.sub(r'\s+Fun.*$', '', m.group(2).strip())
                continue
            if '***' in line and current_nome:
                asterisk_match = re.search(r'\*+([\d*,]+)\*', line)
                if asterisk_match:
                    valor_str = asterisk_match.group(0).replace('*', '').replace(',', '.')
                    try:
                        registros.append({'nome': current_nome, 'nome_norm': normalizar_nome(current_nome),
                            'valor': float(valor_str), 'origem': origem, 'tipo': 'CLT', 'cpf': ''})
                    except (ValueError, TypeError):
                        pass
                current_nome = None
    pdf.close()
    return registros


def detectar_tipo_excel(file_bytes):
    df = pd.read_excel(BytesIO(file_bytes), header=None)
    for i in range(min(10, len(df))):
        row_str = ' '.join(str(v) for v in df.iloc[i].values if pd.notna(v))
        if 'Chave Pix' in row_str or 'Tipo de Pix' in row_str:
            return 'RPA'
    if len(df) > 6 and pd.notna(df.iloc[5][2]) and 'Nome' in str(df.iloc[5][2]):
        return 'RPA'
    return 'CLT'


# ============================================================
# MATCHING
# ============================================================

def _tentar_match_parcial(nome_folha, banco_disponivel):
    nome_v1 = nome_folha.replace('SOUZA', 'SOUSA')
    nome_v2 = nome_folha.replace('SOUSA', 'SOUZA')
    for bidx, brow in banco_disponivel.iterrows():
        nome_banco = brow['nome_norm']
        if len(nome_banco) < 10:
            continue
        if nome_folha.startswith(nome_banco) or nome_banco.startswith(nome_folha[:len(nome_banco)]):
            return (bidx, brow)
        if nome_banco in (nome_v1, nome_v2):
            return (bidx, brow)
        if nome_v1.startswith(nome_banco) or nome_v2.startswith(nome_banco):
            return (bidx, brow)
        if nome_banco.startswith(nome_v1[:len(nome_banco)]) or nome_banco.startswith(nome_v2[:len(nome_banco)]):
            return (bidx, brow)
    return None


def executar_comparacao(df_banco, df_folhas):
    correspondidos = []
    nao_encontrados_banco = []
    divergencia_valor = []
    divergencia_nome = []
    atencao = []
    banco_usado = set()

    for _, folha_row in df_folhas.iterrows():
        nome_folha = folha_row['nome_norm']
        valor_folha = folha_row['valor']
        cpf_folha = folha_row.get('cpf', '')
        banco_disponivel = df_banco[~df_banco.index.isin(banco_usado)]

        match_exato = banco_disponivel[banco_disponivel['nome_norm'] == nome_folha]
        if len(match_exato) > 0:
            brow = match_exato.iloc[0]
            banco_usado.add(match_exato.index[0])
            if abs(brow['valor'] - valor_folha) < 0.02:
                correspondidos.append({'Nome Folha': folha_row['nome'], 'Nome Banco': brow['nome'],
                    'Valor Folha': valor_folha, 'Valor Banco': brow['valor'],
                    'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': 'OK'})
            else:
                divergencia_valor.append({'Nome Folha': folha_row['nome'], 'Nome Banco': brow['nome'],
                    'Valor Folha': valor_folha, 'Valor Banco': brow['valor'],
                    'Diferenca': round(brow['valor'] - valor_folha, 2),
                    'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': 'DIVERGENCIA DE VALOR'})
            continue

        match_parcial = _tentar_match_parcial(nome_folha, banco_disponivel)
        if match_parcial:
            bidx, brow = match_parcial
            banco_usado.add(bidx)
            if abs(brow['valor'] - valor_folha) < 0.02:
                divergencia_nome.append({'Nome Folha': folha_row['nome'], 'Nome Banco': brow['nome'],
                    'Valor Folha': valor_folha, 'Valor Banco': brow['valor'],
                    'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': 'NOME DIFERENTE - VALOR OK'})
            else:
                divergencia_nome.append({'Nome Folha': folha_row['nome'], 'Nome Banco': brow['nome'],
                    'Valor Folha': valor_folha, 'Valor Banco': brow['valor'],
                    'Diferenca': round(brow['valor'] - valor_folha, 2),
                    'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': 'NOME E VALOR DIFERENTES'})
            continue

        if cpf_folha and len(cpf_folha) >= 6:
            digitos_meio = cpf_folha[3:9]
            match_cpf = banco_disponivel[banco_disponivel['cpf_digitos'] == digitos_meio]
            if len(match_cpf) > 0:
                brow = match_cpf.iloc[0]
                banco_usado.add(match_cpf.index[0])
                status = 'MATCH POR CPF - VALOR OK' if abs(brow['valor'] - valor_folha) < 0.02 \
                    else f'MATCH POR CPF - VALOR DIFERENTE'
                atencao.append({'Nome Folha': folha_row['nome'], 'Nome Banco': brow['nome'],
                    'Valor Folha': valor_folha, 'Valor Banco': brow['valor'],
                    'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': status})
                continue

        nao_encontrados_banco.append({'Nome Folha': folha_row['nome'], 'Valor': valor_folha,
            'Origem': folha_row['origem'], 'Tipo': folha_row['tipo'], 'Status': 'NAO ENCONTRADO NO BANCO'})

    banco_sem_folha = df_banco[~df_banco.index.isin(banco_usado)].copy()

    return {
        'correspondidos': correspondidos, 'divergencia_nome': divergencia_nome,
        'divergencia_valor': divergencia_valor, 'atencao': atencao,
        'nao_encontrados_banco': nao_encontrados_banco, 'banco_sem_folha': banco_sem_folha,
    }


# ============================================================
# GERAR EXCEL PARA DOWNLOAD
# ============================================================

def gerar_excel_bytes(resultados, df_banco, df_folhas):
    df_ok = pd.DataFrame(resultados['correspondidos'])
    df_div_nome = pd.DataFrame(resultados['divergencia_nome'])
    df_div_valor = pd.DataFrame(resultados['divergencia_valor'])
    df_atencao = pd.DataFrame(resultados['atencao'])
    df_nao = pd.DataFrame(resultados['nao_encontrados_banco'])
    df_banco_sem = resultados['banco_sem_folha']

    n_clt = len(df_folhas[df_folhas['tipo'] == 'CLT']) if 'tipo' in df_folhas.columns else 0
    n_rpa = len(df_folhas[df_folhas['tipo'] == 'RPA']) if 'tipo' in df_folhas.columns else 0
    v_clt = df_folhas[df_folhas['tipo'] == 'CLT']['valor'].sum() if n_clt > 0 else 0
    v_rpa = df_folhas[df_folhas['tipo'] == 'RPA']['valor'].sum() if n_rpa > 0 else 0
    total_match = len(resultados['correspondidos']) + len(resultados['divergencia_nome'])
    taxa = (total_match / len(df_folhas) * 100) if len(df_folhas) > 0 else 0

    resumo_data = [
        ['Total registros no banco', len(df_banco), f'R$ {df_banco["valor"].sum():,.2f}'],
        ['Total registros nas folhas', len(df_folhas), f'R$ {df_folhas["valor"].sum():,.2f}'],
        ['  - CLT', n_clt, f'R$ {v_clt:,.2f}' if n_clt > 0 else ''],
        ['  - RPA', n_rpa, f'R$ {v_rpa:,.2f}' if n_rpa > 0 else ''],
        ['', '', ''],
        ['Correspondidos OK', len(resultados['correspondidos']),
         f'R$ {sum(r["Valor Folha"] for r in resultados["correspondidos"]):,.2f}' if resultados['correspondidos'] else 'R$ 0,00'],
        ['Divergencia de nome', len(resultados['divergencia_nome']), ''],
        ['Divergencia de valor', len(resultados['divergencia_valor']), ''],
        ['Atencao (CPF)', len(resultados['atencao']), ''],
        ['Nao encontrados no banco', len(resultados['nao_encontrados_banco']), ''],
        ['Banco sem folha', len(df_banco_sem), f'R$ {df_banco_sem["valor"].sum():,.2f}' if len(df_banco_sem) > 0 else 'R$ 0,00'],
        ['', '', ''],
        ['TAXA DE CONCILIACAO', f'{taxa:.1f}%', ''],
    ]
    df_resumo = pd.DataFrame(resumo_data, columns=['Item', 'Quantidade', 'Valor Total'])

    banco_sem_df = pd.DataFrame({
        'Nome': df_banco_sem['nome'].values, 'CPF/CNPJ': df_banco_sem['cpf_cnpj'].values,
        'Referencia': df_banco_sem['ref_empresa'].values, 'Data Pagamento': df_banco_sem['data_pag'].values,
        'Valor': df_banco_sem['valor'].values, 'Status Banco': df_banco_sem['status'].values,
    }) if len(df_banco_sem) > 0 else pd.DataFrame(columns=['Nome', 'CPF/CNPJ', 'Referencia', 'Data Pagamento', 'Valor', 'Status Banco'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumo.to_excel(writer, sheet_name='RESUMO', index=False)
        if len(df_ok) > 0:
            df_ok.to_excel(writer, sheet_name='Correspondidos OK', index=False)
        if len(df_div_nome) > 0:
            df_div_nome.to_excel(writer, sheet_name='Divergencia Nome', index=False)
        if len(df_div_valor) > 0:
            df_div_valor.to_excel(writer, sheet_name='Divergencia Valor', index=False)
        if len(df_atencao) > 0:
            df_atencao.to_excel(writer, sheet_name='Atencao', index=False)
        if len(df_nao) > 0:
            df_nao.to_excel(writer, sheet_name='Nao Encontrados', index=False)
        banco_sem_df.to_excel(writer, sheet_name='Banco sem Folha', index=False)

    # Formatar
    output.seek(0)
    wb = load_workbook(output)
    hfill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    ok_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    err_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    border = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
                    top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.freeze_panes = 'A2'
        for ci, col in enumerate(ws.columns, 1):
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(max(ml + 3, 10), 45)
            for cell in col:
                cell.border = border

        # Cor de fundo por aba
        sheet_name = ws.title
        if sheet_name == 'Correspondidos OK':
            fill = ok_fill
        elif 'Divergencia' in sheet_name:
            fill = warn_fill
        elif sheet_name in ('Atencao', 'Nao Encontrados'):
            fill = err_fill
        else:
            fill = None

        if fill:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.fill = fill

    out2 = BytesIO()
    wb.save(out2)
    out2.seek(0)
    return out2.getvalue()


# ============================================================
# PAGINA PRINCIPAL
# ============================================================

st.title('Comparador Folha x Banco - Staff Force')

st.markdown("""
<style>
    [data-testid="stDownloadButton"] button {
        background-color: #2E7D32 !important;
        color: white !important;
        border: none !important;
    }
    [data-testid="stDownloadButton"] button:hover {
        background-color: #1B5E20 !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

st.markdown('---')

# --- UPLOAD ---
col_banco, col_folhas = st.columns(2)

with col_banco:
    st.subheader('Arquivo do Banco')
    banco_file = st.file_uploader(
        'ConsultaPagamentos (.xls / .xlsx)',
        type=['xls', 'xlsx'],
        key='banco',
    )

with col_folhas:
    st.subheader('Folhas de Pagamento')
    st.caption('Excel (.xlsx) ou PDF (.pdf) - varios arquivos')
    folha_files = st.file_uploader(
        'Selecione as folhas',
        type=['xlsx', 'xls', 'pdf'],
        accept_multiple_files=True,
        key='folhas',
    )

# Configurar origens
folha_configs = []
if folha_files:
    st.markdown('**Origens das folhas:**')
    cols = st.columns(min(len(folha_files), 4))
    for i, f in enumerate(folha_files):
        with cols[i % len(cols)]:
            origem = st.text_input(f.name[:30], value=sugerir_origem(f.name), key=f'orig_{i}')
            folha_configs.append({'file': f, 'origem': origem})

st.markdown('---')

# --- BOTAO EXECUTAR ---
col_btn, col_space = st.columns([1, 3])
with col_btn:
    executar = st.button('EXECUTAR COMPARACAO', type='primary', use_container_width=True)

if executar:
    if not banco_file:
        st.error('Selecione o arquivo do banco!')
        st.stop()
    if not folha_files:
        st.error('Selecione pelo menos uma folha!')
        st.stop()

    with st.spinner('Processando...'):
        df_banco = ler_banco_bytes(banco_file.read(), banco_file.name)

        todos_registros = []
        for cfg in folha_configs:
            f = cfg['file']
            fb = f.read()
            if f.name.lower().endswith('.pdf'):
                regs = ler_folha_pdf_bytes(fb, cfg['origem'])
            else:
                tipo = detectar_tipo_excel(fb)
                if tipo == 'RPA':
                    regs = ler_folha_rpa_excel_bytes(fb, cfg['origem'])
                else:
                    regs = ler_folha_clt_bytes(fb, cfg['origem'])
            todos_registros.extend(regs)

        if not todos_registros:
            st.error('Nenhum registro encontrado nas folhas!')
            st.stop()

        df_folhas = pd.DataFrame(todos_registros)
        resultados = executar_comparacao(df_banco, df_folhas)

        st.session_state['resultados'] = resultados
        st.session_state['df_banco'] = df_banco
        st.session_state['df_folhas'] = df_folhas
        st.session_state['processado'] = True

# --- RESULTADOS ---
if st.session_state.get('processado'):
    resultados = st.session_state['resultados']
    df_banco = st.session_state['df_banco']
    df_folhas = st.session_state['df_folhas']

    n_ok = len(resultados['correspondidos'])
    n_div_nome = len(resultados['divergencia_nome'])
    n_div_valor = len(resultados['divergencia_valor'])
    n_atencao = len(resultados['atencao'])
    n_nao = len(resultados['nao_encontrados_banco'])
    n_banco_sem = len(resultados['banco_sem_folha'])
    total_match = n_ok + n_div_nome
    taxa = (total_match / len(df_folhas) * 100) if len(df_folhas) > 0 else 0

    # BOTAO EXPORTAR EXCEL - BEM VISIVEL
    st.markdown('---')
    excel_bytes = gerar_excel_bytes(resultados, df_banco, df_folhas)
    col_dl, col_info = st.columns([1, 2])
    with col_dl:
        st.download_button(
            label='EXPORTAR PARA EXCEL',
            data=excel_bytes,
            file_name=f'Resultado_Comparacao_{date.today().strftime("%d.%m.%y")}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary',
            use_container_width=True,
        )
    with col_info:
        st.info(f'Banco: {len(df_banco)} registros | Folhas: {len(df_folhas)} registros | '
                f'Conciliados: {total_match} ({taxa:.1f}%) | Nao encontrados: {n_nao}')

    # --- TABELAS ---
    st.markdown('---')

    # CORRESPONDIDOS OK
    st.subheader(f'Correspondidos OK ({n_ok})')
    if n_ok > 0:
        df_ok = pd.DataFrame(resultados['correspondidos'])
        busca_ok = st.text_input('Buscar por nome:', key='b_ok')
        d = df_ok.copy()
        if busca_ok:
            d = d[d.apply(lambda r: busca_ok.lower() in str(r).lower(), axis=1)]
        st.dataframe(d, use_container_width=True, height=min(35 * len(d) + 40, 400))
    else:
        st.warning('Nenhum correspondido')

    # DIVERGENCIA DE NOME
    if n_div_nome > 0:
        st.markdown('---')
        st.subheader(f'Divergencia de Nome ({n_div_nome})')
        df_dn = pd.DataFrame(resultados['divergencia_nome'])
        st.dataframe(df_dn, use_container_width=True, height=min(35 * len(df_dn) + 40, 400))

    # DIVERGENCIA DE VALOR
    if n_div_valor > 0:
        st.markdown('---')
        st.subheader(f'Divergencia de Valor ({n_div_valor})')
        df_dv = pd.DataFrame(resultados['divergencia_valor'])
        st.dataframe(df_dv, use_container_width=True, height=min(35 * len(df_dv) + 40, 400))

    # ATENCAO (CPF match)
    if n_atencao > 0:
        st.markdown('---')
        st.subheader(f'Atencao - Match por CPF ({n_atencao})')
        df_at = pd.DataFrame(resultados['atencao'])
        st.dataframe(df_at, use_container_width=True, height=min(35 * len(df_at) + 40, 300))

    # NAO ENCONTRADOS NO BANCO
    if n_nao > 0:
        st.markdown('---')
        st.subheader(f'NAO ENCONTRADOS NO BANCO ({n_nao})')
        df_nao = pd.DataFrame(resultados['nao_encontrados_banco'])
        st.dataframe(df_nao, use_container_width=True, height=min(35 * len(df_nao) + 40, 300))

    # BANCO SEM FOLHA
    if n_banco_sem > 0:
        st.markdown('---')
        st.subheader(f'No Banco mas nao nas Folhas ({n_banco_sem})')
        df_bs = resultados['banco_sem_folha']
        df_bs_display = pd.DataFrame({
            'Nome': df_bs['nome'].values, 'CPF/CNPJ': df_bs['cpf_cnpj'].values,
            'Referencia': df_bs['ref_empresa'].values, 'Valor': df_bs['valor'].values,
            'Status': df_bs['status'].values,
        })
        busca_bs = st.text_input('Buscar por nome:', key='b_bs')
        d = df_bs_display.copy()
        if busca_bs:
            d = d[d.apply(lambda r: busca_bs.lower() in str(r).lower(), axis=1)]
        st.dataframe(d, use_container_width=True, height=min(35 * len(d) + 40, 400))

    st.markdown('---')
    st.caption('Staff Force - Comparador Folha x Banco')
