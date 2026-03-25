import pandas as pd
import unicodedata
import re
import glob
import os
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# FUNCOES UTILITARIAS
# ============================================================

def normalizar_nome(nome):
    if pd.isna(nome):
        return ''
    nome = str(nome).strip().upper()
    nome = unicodedata.normalize('NFD', nome)
    nome = ''.join(c for c in nome if unicodedata.category(c) != 'Mn')
    nome = re.sub(r'\s+', ' ', nome)
    # Strip CNPJ prefixado (ex: "59.128.455 ERICLES..." -> "ERICLES...")
    nome = re.sub(r'^\d{2}\.\d{3}\.\d{3}\s+', '', nome)
    return nome


def extrair_digitos_visiveis_cpf(cpf_mascarado):
    """Extrai digitos visiveis de CPF mascarado: ***.716.624-** -> 716624"""
    if pd.isna(cpf_mascarado):
        return ''
    cpf_str = str(cpf_mascarado)
    digitos = re.findall(r'\d+', cpf_str)
    return ''.join(digitos)


def cpf_limpo(cpf):
    """Remove formatacao do CPF: 756.716.624-00 -> 75671662400"""
    if pd.isna(cpf):
        return ''
    return re.sub(r'[.\-\s]', '', str(cpf).strip())


def extrair_data_arquivo(nome_arquivo):
    """Extrai data do nome do arquivo banco (ex: 25.03.26)"""
    match = re.search(r'(\d{2}\.\d{2}\.\d{2})', nome_arquivo)
    if match:
        return match.group(1)
    hoje = date.today()
    return hoje.strftime('%d.%m.%y')


def derivar_origem_clt(nome_arquivo):
    """Deriva label de origem do nome do arquivo CLT"""
    base = os.path.basename(nome_arquivo)
    # Extrair codigo (ex: 0060, 0041)
    cod_match = re.match(r'(\d{4})', base)
    codigo = cod_match.group(1) if cod_match else ''

    # Extrair local (entre LIQUIDO_PG_ e " - ")
    local_match = re.search(r'LIQUIDO_PG_(\w+)', base, re.IGNORECASE)
    local = local_match.group(1) if local_match else ''

    # Verificar se e quinzenal
    is_quinz = 'QUINZ' in base.upper()
    # Se local ja e QUINZ, pegar proxima parte
    if local.upper() == 'QUINZ':
        after_match = re.search(r'LIQUIDO_PG_QUINZ[_ -]+(?:OP\s+)?(\w+)', base, re.IGNORECASE)
        local = after_match.group(1) if after_match else local
        return f'QUINZ {local} {codigo}'.strip()

    if is_quinz:
        return f'QUINZ {local} {codigo}'.strip()

    return f'{local} {codigo}'.strip()


# ============================================================
# AUTO-DETECCAO DE ARQUIVOS
# ============================================================

def auto_detectar_arquivos(pasta='.'):
    """Detecta automaticamente os arquivos na pasta"""
    config = {
        'arquivo_banco': None,
        'arquivos_clt': [],
        'arquivos_rpa': [],
        'data_referencia': None,
    }

    todos = os.listdir(pasta)

    # Banco: ConsultaPagamentos*.xls*
    bancos = [f for f in todos if f.lower().startswith('consultapagamentos') and f.lower().endswith(('.xls', '.xlsx'))]
    # Pegar o mais recente (maior data ou ultimo modificado)
    if bancos:
        bancos.sort(key=lambda f: os.path.getmtime(os.path.join(pasta, f)), reverse=True)
        config['arquivo_banco'] = bancos[0]
        config['data_referencia'] = extrair_data_arquivo(bancos[0])

    # CLT: *LIQUIDO_PG*.xlsx
    clts = [f for f in todos if 'LIQUIDO_PG' in f.upper() and f.lower().endswith('.xlsx')]
    for arq in sorted(clts):
        origem = derivar_origem_clt(arq)
        config['arquivos_clt'].append((arq, origem))

    # RPA: *RPA*.xlsx (excluir resultados)
    rpas = [f for f in todos if 'RPA' in f.upper() and f.lower().endswith('.xlsx')
            and 'resultado' not in f.lower()]
    for arq in sorted(rpas):
        config['arquivos_rpa'].append((arq, 'RPA'))

    if not config['data_referencia']:
        config['data_referencia'] = date.today().strftime('%d.%m.%y')

    return config


def solicitar_configuracao():
    """Interface interativa para confirmar/ajustar arquivos"""
    config = auto_detectar_arquivos()

    print('=' * 64)
    print('  COMPARADOR FOLHA x BANCO - Staff Force')
    print('=' * 64)
    print()
    print('Arquivos encontrados na pasta:')

    if config['arquivo_banco']:
        print(f'  [BANCO] {config["arquivo_banco"]}')
    else:
        print('  [BANCO] Nenhum arquivo encontrado!')

    for arq, origem in config['arquivos_clt']:
        print(f'  [CLT]   {arq}  ({origem})')

    for arq, origem in config['arquivos_rpa']:
        print(f'  [RPA]   {arq}')

    if not config['arquivos_clt'] and not config['arquivos_rpa']:
        print('  [FOLHA] Nenhum arquivo de folha encontrado!')

    print()

    resposta = input('Usar estes arquivos? (S/n): ').strip().lower()
    if resposta == 'n':
        # Config manual
        banco = input('Arquivo do banco (ConsultaPagamentos): ').strip()
        if banco:
            config['arquivo_banco'] = banco
            config['data_referencia'] = extrair_data_arquivo(banco)

        print('Arquivos CLT (um por linha, vazio para terminar):')
        config['arquivos_clt'] = []
        while True:
            arq = input('  Arquivo CLT: ').strip()
            if not arq:
                break
            origem = input('  Origem (ex: RJ 0060): ').strip()
            config['arquivos_clt'].append((arq, origem))

        print('Arquivos RPA (um por linha, vazio para terminar):')
        config['arquivos_rpa'] = []
        while True:
            arq = input('  Arquivo RPA: ').strip()
            if not arq:
                break
            config['arquivos_rpa'].append((arq, 'RPA'))

    print()
    return config


# ============================================================
# FUNCOES DE LEITURA
# ============================================================

def ler_banco(arquivo):
    """Le arquivo de lancamentos bancarios"""
    df_raw = pd.read_excel(arquivo, header=None)

    # Buscar header dinamicamente (row com "favorecido")
    header_row = None
    for i in range(min(30, len(df_raw))):
        row_str = ' '.join(str(v).lower() for v in df_raw.iloc[i].values if pd.notna(v))
        if 'favorecido' in row_str or 'benefici' in row_str:
            header_row = i
            break

    if header_row is None:
        header_row = 18  # fallback
        print(f'  AVISO: Header do banco nao encontrado automaticamente, usando linha {header_row}')

    df_banco = df_raw.iloc[header_row + 1:].copy()
    df_banco.columns = ['nome', 'cpf_cnpj', 'tipo_pag', 'ref_empresa', 'data_pag', 'valor', 'status']
    df_banco = df_banco.dropna(subset=['nome'])
    df_banco = df_banco[df_banco['nome'].astype(str).str.strip() != '']
    df_banco = df_banco[df_banco['nome'] != 'Total:']
    df_banco = df_banco[~df_banco['nome'].astype(str).str.lower().str.contains('favorecido')]
    df_banco['valor'] = pd.to_numeric(df_banco['valor'], errors='coerce')
    df_banco['nome_norm'] = df_banco['nome'].apply(normalizar_nome)
    df_banco['cpf_digitos'] = df_banco['cpf_cnpj'].apply(extrair_digitos_visiveis_cpf)

    print(f'  Banco: {len(df_banco)} registros | R$ {df_banco["valor"].sum():,.2f}')
    return df_banco


def ler_folha_clt(arquivo, origem):
    """Le uma folha CLT (Liquido Geral)"""
    df = pd.read_excel(arquivo, header=None)

    # Dados a partir da linha 8 (padrao)
    dados = df.iloc[8:].copy()
    dados = dados[dados[0].notna()]
    dados = dados[~dados[0].astype(str).str.contains('Total|Dpto|TOTAL|Resumo', na=False)]

    registros = []
    for _, row in dados.iterrows():
        nome = str(row[0]).strip()
        valor = row[15] if pd.notna(row[15]) else (row[16] if len(row) > 16 and pd.notna(row[16]) else None)
        if nome and valor is not None:
            try:
                registros.append({
                    'nome': nome,
                    'nome_norm': normalizar_nome(nome),
                    'valor': float(valor),
                    'origem': origem,
                    'tipo': 'CLT',
                    'cpf': '',
                })
            except (ValueError, TypeError):
                pass

    print(f'  CLT {origem}: {len(registros)} registros')
    return registros


def ler_folha_rpa(arquivo, origem='RPA'):
    """Le folha RPA"""
    df = pd.read_excel(arquivo, header=None)

    # Header na row 5, dados a partir da row 6
    dados = df.iloc[6:].copy()
    dados = dados[dados[2].notna()]
    dados = dados[~dados[2].astype(str).str.contains('Total|Nome|TOTAL', na=False)]

    registros = []
    for _, row in dados.iterrows():
        nome = str(row[2]).strip()
        valor = row[3]
        cpf = str(row[10]).strip() if len(row) > 10 and pd.notna(row[10]) else ''
        if nome and pd.notna(valor):
            try:
                registros.append({
                    'nome': nome,
                    'nome_norm': normalizar_nome(nome),
                    'valor': float(valor),
                    'origem': origem,
                    'tipo': 'RPA',
                    'cpf': cpf_limpo(cpf),
                })
            except (ValueError, TypeError):
                pass

    print(f'  RPA: {len(registros)} registros')
    return registros


def carregar_todas_folhas(config):
    """Carrega CLT + RPA em um unico DataFrame"""
    todos = []

    for arq, origem in config['arquivos_clt']:
        todos.extend(ler_folha_clt(arq, origem))

    for arq, origem in config['arquivos_rpa']:
        todos.extend(ler_folha_rpa(arq, origem))

    df = pd.DataFrame(todos)
    if len(df) == 0:
        print('  AVISO: Nenhum registro encontrado nas folhas!')
    return df


# ============================================================
# MATCHING
# ============================================================

def executar_comparacao(df_banco, df_folhas):
    """Compara folhas vs banco e retorna resultados classificados"""
    correspondidos = []
    nao_encontrados_banco = []
    divergencia_valor = []
    divergencia_nome = []
    atencao = []
    banco_usado = set()

    for _, folha_row in df_folhas.iterrows():
        nome_folha = folha_row['nome_norm']
        valor_folha = folha_row['valor']
        cpf_folha = folha_row.get('cpf', '')

        banco_disponivel = df_banco[~df_banco.index.isin(banco_usado)]

        # === Nivel 1: Match exato por nome normalizado ===
        match_exato = banco_disponivel[banco_disponivel['nome_norm'] == nome_folha]

        if len(match_exato) > 0:
            brow = match_exato.iloc[0]
            bidx = match_exato.index[0]
            banco_usado.add(bidx)
            _classificar(folha_row, brow, correspondidos, divergencia_valor, match_tipo='EXATO')
            continue

        # === Nivel 2: Match parcial (truncamento + SOUZA/SOUSA) ===
        match_parcial = _tentar_match_parcial(nome_folha, banco_disponivel)

        if match_parcial:
            bidx, brow = match_parcial
            banco_usado.add(bidx)

            if abs(brow['valor'] - valor_folha) < 0.02:
                divergencia_nome.append({
                    'nome_folha': folha_row['nome'],
                    'nome_banco': brow['nome'],
                    'valor_folha': valor_folha,
                    'valor_banco': brow['valor'],
                    'diferenca': brow['valor'] - valor_folha,
                    'origem': folha_row['origem'],
                    'tipo': folha_row['tipo'],
                    'status': 'NOME DIFERENTE - VALOR OK',
                })
            else:
                divergencia_nome.append({
                    'nome_folha': folha_row['nome'],
                    'nome_banco': brow['nome'],
                    'valor_folha': valor_folha,
                    'valor_banco': brow['valor'],
                    'diferenca': brow['valor'] - valor_folha,
                    'origem': folha_row['origem'],
                    'tipo': folha_row['tipo'],
                    'status': f'NOME DIFERENTE - DIVERGENCIA VALOR R$ {brow["valor"] - valor_folha:.2f}',
                })
            continue

        # === Nivel 3: Match por CPF (so para RPA com CPF) ===
        if cpf_folha and len(cpf_folha) >= 6:
            digitos_meio = cpf_folha[3:9]  # 6 digitos do meio
            match_cpf = banco_disponivel[
                banco_disponivel['cpf_digitos'] == digitos_meio
            ]
            if len(match_cpf) > 0:
                brow = match_cpf.iloc[0]
                bidx = match_cpf.index[0]
                banco_usado.add(bidx)

                if abs(brow['valor'] - valor_folha) < 0.02:
                    atencao.append({
                        'nome_folha': folha_row['nome'],
                        'nome_banco': brow['nome'],
                        'valor_folha': valor_folha,
                        'valor_banco': brow['valor'],
                        'origem': folha_row['origem'],
                        'tipo': folha_row['tipo'],
                        'status': f'MATCH POR CPF - nomes diferentes, mesmo valor',
                    })
                else:
                    atencao.append({
                        'nome_folha': folha_row['nome'],
                        'nome_banco': brow['nome'],
                        'valor_folha': valor_folha,
                        'valor_banco': brow['valor'],
                        'origem': folha_row['origem'],
                        'tipo': folha_row['tipo'],
                        'status': f'MATCH POR CPF - DIVERGENCIA VALOR R$ {brow["valor"] - valor_folha:.2f}',
                    })
                continue

        # === Nao encontrado ===
        nao_encontrados_banco.append({
            'nome': folha_row['nome'],
            'valor': valor_folha,
            'origem': folha_row['origem'],
            'tipo': folha_row['tipo'],
            'status': 'NAO ENCONTRADO NO BANCO',
        })

    banco_sem_folha = df_banco[~df_banco.index.isin(banco_usado)].copy()

    return {
        'correspondidos': correspondidos,
        'divergencia_nome': divergencia_nome,
        'divergencia_valor': divergencia_valor,
        'atencao': atencao,
        'nao_encontrados_banco': nao_encontrados_banco,
        'banco_sem_folha': banco_sem_folha,
    }


def _tentar_match_parcial(nome_folha, banco_disponivel):
    """Tenta match por truncamento e variacao SOUZA/SOUSA"""
    nome_v1 = nome_folha.replace('SOUZA', 'SOUSA')
    nome_v2 = nome_folha.replace('SOUSA', 'SOUZA')

    for bidx, brow in banco_disponivel.iterrows():
        nome_banco = brow['nome_norm']

        if len(nome_banco) < 10:
            continue

        # Truncamento simples
        if nome_folha.startswith(nome_banco) or nome_banco.startswith(nome_folha[:len(nome_banco)]):
            return (bidx, brow)

        # SOUZA/SOUSA exato
        if nome_banco in (nome_v1, nome_v2):
            return (bidx, brow)

        # Truncamento + SOUZA/SOUSA
        if nome_v1.startswith(nome_banco) or nome_v2.startswith(nome_banco):
            return (bidx, brow)

        if nome_banco.startswith(nome_v1[:len(nome_banco)]) or nome_banco.startswith(nome_v2[:len(nome_banco)]):
            return (bidx, brow)

    return None


def _classificar(folha_row, brow, correspondidos, divergencia_valor, match_tipo='EXATO'):
    """Classifica match como OK ou divergencia de valor"""
    valor_folha = folha_row['valor']
    if abs(brow['valor'] - valor_folha) < 0.02:
        correspondidos.append({
            'nome_folha': folha_row['nome'],
            'nome_banco': brow['nome'],
            'valor_folha': valor_folha,
            'valor_banco': brow['valor'],
            'diferenca': brow['valor'] - valor_folha,
            'origem': folha_row['origem'],
            'tipo': folha_row['tipo'],
            'status': 'OK',
        })
    else:
        divergencia_valor.append({
            'nome_folha': folha_row['nome'],
            'nome_banco': brow['nome'],
            'valor_folha': valor_folha,
            'valor_banco': brow['valor'],
            'diferenca': brow['valor'] - valor_folha,
            'origem': folha_row['origem'],
            'tipo': folha_row['tipo'],
            'status': f'DIVERGENCIA VALOR R$ {brow["valor"] - valor_folha:.2f}',
        })


# ============================================================
# SAIDA EXCEL
# ============================================================

def gerar_excel(resultados, config, df_banco, df_folhas):
    """Gera arquivo Excel formatado com 5 abas"""
    data_ref = config['data_referencia']
    nome_arquivo = f'Resultado_Comparacao_{data_ref}.xlsx'

    # Preparar DataFrames
    df_ok = pd.DataFrame(resultados['correspondidos'])
    df_div_nome = pd.DataFrame(resultados['divergencia_nome'])
    df_div_valor = pd.DataFrame(resultados['divergencia_valor'])
    df_atencao_list = resultados['atencao'] + resultados['nao_encontrados_banco']
    df_atencao = pd.DataFrame(df_atencao_list) if df_atencao_list else pd.DataFrame()
    df_banco_sem = resultados['banco_sem_folha']

    # RESUMO
    n_clt = len(df_folhas[df_folhas['tipo'] == 'CLT']) if 'tipo' in df_folhas.columns else 0
    n_rpa = len(df_folhas[df_folhas['tipo'] == 'RPA']) if 'tipo' in df_folhas.columns else 0
    v_clt = df_folhas[df_folhas['tipo'] == 'CLT']['valor'].sum() if n_clt > 0 else 0
    v_rpa = df_folhas[df_folhas['tipo'] == 'RPA']['valor'].sum() if n_rpa > 0 else 0

    resumo_data = [
        ['Total registros no banco', len(df_banco), f'R$ {df_banco["valor"].sum():,.2f}'],
        ['Total registros nas folhas', len(df_folhas), f'R$ {df_folhas["valor"].sum():,.2f}'],
        ['  - CLT', n_clt, f'R$ {v_clt:,.2f}' if n_clt > 0 else ''],
        ['  - RPA', n_rpa, f'R$ {v_rpa:,.2f}' if n_rpa > 0 else ''],
    ]

    # Breakdown por origem
    if 'origem' in df_folhas.columns:
        for origem in df_folhas['origem'].unique():
            sub = df_folhas[df_folhas['origem'] == origem]
            resumo_data.append([f'    . {origem}', len(sub), f'R$ {sub["valor"].sum():,.2f}'])

    resumo_data.extend([
        ['', '', ''],
        ['Correspondidos OK', len(resultados['correspondidos']),
         f'R$ {sum(r["valor_folha"] for r in resultados["correspondidos"]):,.2f}' if resultados['correspondidos'] else 'R$ 0,00'],
        ['Divergencia de nome', len(resultados['divergencia_nome']), ''],
        ['Divergencia de valor', len(resultados['divergencia_valor']), ''],
        ['Atencao (CPF match / nao encontrados)', len(df_atencao_list), ''],
        ['Banco sem folha', len(df_banco_sem), f'R$ {df_banco_sem["valor"].sum():,.2f}' if len(df_banco_sem) > 0 else 'R$ 0,00'],
        ['', '', ''],
    ])

    total_folhas = len(df_folhas)
    total_match = len(resultados['correspondidos']) + len(resultados['divergencia_nome'])
    taxa = (total_match / total_folhas * 100) if total_folhas > 0 else 0
    resumo_data.append(['TAXA DE CONCILIACAO', f'{taxa:.1f}%', ''])

    df_resumo = pd.DataFrame(resumo_data, columns=['Item', 'Quantidade', 'Valor Total'])

    # Preparar banco sem folha
    banco_sem_cols = {}
    if len(df_banco_sem) > 0:
        banco_sem_cols = pd.DataFrame({
            'Nome': df_banco_sem['nome'].values,
            'CPF/CNPJ': df_banco_sem['cpf_cnpj'].values,
            'Referencia': df_banco_sem['ref_empresa'].values,
            'Data Pagamento': df_banco_sem['data_pag'].values,
            'Valor': df_banco_sem['valor'].values,
            'Status Banco': df_banco_sem['status'].values,
        })
    else:
        banco_sem_cols = pd.DataFrame(columns=['Nome', 'CPF/CNPJ', 'Referencia', 'Data Pagamento', 'Valor', 'Status Banco'])

    # Escrever Excel
    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        df_resumo.to_excel(writer, sheet_name='RESUMO', index=False)

        if len(df_ok) > 0:
            cols_ok = ['nome_folha', 'nome_banco', 'valor_folha', 'valor_banco', 'diferenca', 'origem', 'tipo', 'status']
            df_ok_out = df_ok[cols_ok].copy()
            df_ok_out.columns = ['Nome Folha', 'Nome Banco', 'Valor Folha', 'Valor Banco', 'Diferenca', 'Origem', 'Tipo', 'Status']
            df_ok_out.to_excel(writer, sheet_name='Correspondidos OK', index=False)
        else:
            pd.DataFrame(columns=['Nome Folha', 'Nome Banco', 'Valor Folha', 'Valor Banco', 'Diferenca', 'Origem', 'Tipo', 'Status']).to_excel(
                writer, sheet_name='Correspondidos OK', index=False)

        if len(df_div_nome) > 0:
            cols_dn = ['nome_folha', 'nome_banco', 'valor_folha', 'valor_banco', 'diferenca', 'origem', 'tipo', 'status']
            df_dn_out = df_div_nome[cols_dn].copy()
            df_dn_out.columns = ['Nome Folha', 'Nome Banco', 'Valor Folha', 'Valor Banco', 'Diferenca', 'Origem', 'Tipo', 'Status']
            df_dn_out.to_excel(writer, sheet_name='Divergencia Nome', index=False)
        else:
            pd.DataFrame(columns=['Nome Folha', 'Nome Banco', 'Valor Folha', 'Valor Banco', 'Diferenca', 'Origem', 'Tipo', 'Status']).to_excel(
                writer, sheet_name='Divergencia Nome', index=False)

        if len(df_atencao) > 0:
            df_atencao.to_excel(writer, sheet_name='Atencao', index=False)
        else:
            pd.DataFrame(columns=['Nome', 'Valor', 'Origem', 'Tipo', 'Status']).to_excel(
                writer, sheet_name='Atencao', index=False)

        banco_sem_cols.to_excel(writer, sheet_name='Banco sem Folha', index=False)

    # Aplicar formatacao
    _formatar_excel(nome_arquivo)

    print(f'\nArquivo salvo: {nome_arquivo}')
    return nome_arquivo


def _formatar_excel(nome_arquivo):
    """Aplica formatacao visual ao Excel"""
    wb = load_workbook(nome_arquivo)

    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ok_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    error_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    for ws in wb.worksheets:
        # Formatar headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Freeze panes
        ws.freeze_panes = 'A2'

        # Auto-ajustar largura
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                cell.border = thin_border
                try:
                    val = str(cell.value) if cell.value else ''
                    max_len = max(max_len, len(val))
                except:
                    pass
            adjusted = min(max_len + 3, 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 10)

        # Formatar colunas de valor
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                header_val = ws.cell(row=1, column=cell.column).value
                if header_val and any(kw in str(header_val).lower() for kw in ['valor', 'diferenca', 'total']):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

    # Cores condicionais na aba RESUMO
    ws_resumo = wb['RESUMO']
    for row in ws_resumo.iter_rows(min_row=2, max_row=ws_resumo.max_row):
        item = str(row[0].value or '').lower()
        if 'correspondidos' in item:
            for cell in row:
                cell.fill = ok_fill
        elif 'divergencia' in item or 'atencao' in item:
            for cell in row:
                cell.fill = warn_fill
        elif 'taxa' in item:
            for cell in row:
                cell.fill = ok_fill
                cell.font = Font(bold=True, size=12)

    # Cores na aba Correspondidos OK
    if 'Correspondidos OK' in wb.sheetnames:
        ws_ok = wb['Correspondidos OK']
        for row in ws_ok.iter_rows(min_row=2, max_row=ws_ok.max_row):
            for cell in row:
                cell.fill = ok_fill

    # Cores na aba Divergencia Nome
    if 'Divergencia Nome' in wb.sheetnames:
        ws_dn = wb['Divergencia Nome']
        for row in ws_dn.iter_rows(min_row=2, max_row=ws_dn.max_row):
            for cell in row:
                cell.fill = warn_fill

    # Cores na aba Atencao
    if 'Atencao' in wb.sheetnames:
        ws_at = wb['Atencao']
        for row in ws_at.iter_rows(min_row=2, max_row=ws_at.max_row):
            status = str(row[-1].value or '').lower()
            fill = error_fill if 'nao encontrado' in status else warn_fill
            for cell in row:
                cell.fill = fill

    wb.save(nome_arquivo)


# ============================================================
# SAIDA CONSOLE
# ============================================================

def imprimir_resumo(resultados, df_banco, df_folhas):
    """Imprime resumo no console"""
    print()
    print('=' * 64)
    print('  RESUMO DA COMPARACAO')
    print('=' * 64)
    print(f'  Registros no banco:          {len(df_banco)}')
    print(f'  Registros nas folhas:        {len(df_folhas)}')

    if 'tipo' in df_folhas.columns:
        n_clt = len(df_folhas[df_folhas['tipo'] == 'CLT'])
        n_rpa = len(df_folhas[df_folhas['tipo'] == 'RPA'])
        if n_clt > 0:
            print(f'    - CLT:                     {n_clt}')
        if n_rpa > 0:
            print(f'    - RPA:                     {n_rpa}')

    print(f'  Correspondidos OK:           {len(resultados["correspondidos"])}')
    print(f'  Divergencia de nome:         {len(resultados["divergencia_nome"])}')
    print(f'  Divergencia de valor:        {len(resultados["divergencia_valor"])}')
    print(f'  Atencao (CPF / outros):      {len(resultados["atencao"])}')
    print(f'  Nao encontrados no banco:    {len(resultados["nao_encontrados_banco"])}')
    print(f'  Banco sem folha:             {len(resultados["banco_sem_folha"])}')

    total_folhas = len(df_folhas)
    total_match = len(resultados['correspondidos']) + len(resultados['divergencia_nome'])
    taxa = (total_match / total_folhas * 100) if total_folhas > 0 else 0
    print(f'  TAXA DE CONCILIACAO:         {taxa:.1f}%')
    print('=' * 64)

    if resultados['nao_encontrados_banco']:
        print()
        print('  *** NAO ENCONTRADOS NO BANCO ***')
        for r in resultados['nao_encontrados_banco']:
            print(f'    - {r["nome"]} | R$ {r["valor"]:.2f} | {r["origem"]}')

    if resultados['atencao']:
        print()
        print('  *** ATENCAO ***')
        for r in resultados['atencao']:
            print(f'    - {r["nome_folha"]} -> {r.get("nome_banco", "?")} | {r["status"]}')


# ============================================================
# MAIN
# ============================================================

def main():
    config = solicitar_configuracao()

    if not config['arquivo_banco']:
        print('ERRO: Nenhum arquivo do banco selecionado!')
        return

    if not config['arquivos_clt'] and not config['arquivos_rpa']:
        print('ERRO: Nenhuma folha de pagamento selecionada!')
        return

    print('Carregando dados...')
    df_banco = ler_banco(config['arquivo_banco'])
    df_folhas = carregar_todas_folhas(config)

    if len(df_folhas) == 0:
        print('ERRO: Nenhum registro encontrado nas folhas!')
        return

    print('\nExecutando comparacao...')
    resultados = executar_comparacao(df_banco, df_folhas)

    imprimir_resumo(resultados, df_banco, df_folhas)

    print('\nGerando Excel...')
    nome_excel = gerar_excel(resultados, config, df_banco, df_folhas)

    print(f'\nConcluido! Resultado salvo em: {nome_excel}')
    input('\nPressione ENTER para fechar...')


if __name__ == '__main__':
    main()
